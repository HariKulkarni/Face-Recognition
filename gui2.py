import calendar
import streamlit as st
import cv2
import numpy as np
import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side

SUBJECTS = ["DS", "DM", "PHP", "AI"]
NUM_DAYS = 31
MONTH_NAME = "May"
NAME_FILE = "id_name.txt"
EXCEL_FILE = 'attendance_excel/attendance.xlsx'
EXCEL_BASE_DIR = "attendance_excel"  # Base directory for monthly Excel files

# ------------------- Helper Functions -------------------

def initialize_sheet(sheet):
    """Initialize the Excel sheet with proper headers"""
    try:
        # Clear existing content
        sheet.delete_rows(1, sheet.max_row)
        
        # Set headers
        sheet.cell(row=1, column=1, value="ID")
        sheet.cell(row=1, column=2, value="Name")
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        col = 3
        # Date headers
        for day in range(1, NUM_DAYS + 1):
            if col + len(SUBJECTS) - 1 <= sheet.max_column or sheet.max_column == 1:
                sheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(SUBJECTS) - 1)
            sheet.cell(row=1, column=col, value=f"{day} {MONTH_NAME}")
            
            # Subject headers for each day
            for i, subject in enumerate(SUBJECTS):
                sheet.cell(row=2, column=col + i, value=subject)
            col += len(SUBJECTS)

        # Total columns for each subject
        for subject in SUBJECTS:
            sheet.cell(row=1, column=col, value=f"Total {subject}")
            sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            col += 1

        # Summary columns
        sheet.cell(row=1, column=col, value="Total Attended Classes")
        sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

        sheet.cell(row=1, column=col, value="Total Classes")
        sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

        sheet.cell(row=1, column=col, value="Percentage")
        sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

        # Apply alignment
        for row in sheet.iter_rows(min_row=1, max_row=2):
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
        print("Sheet initialized successfully")
        return True
    except Exception as e:
        print(f"Error initializing sheet: {e}")
        return False

def get_column_for_subject(day, subject):
    """Get the column number for a specific day and subject"""
    if day < 1 or day > NUM_DAYS:
        raise ValueError(f"Day must be between 1 and {NUM_DAYS}")
    if subject not in SUBJECTS:
        raise ValueError(f"Subject must be one of {SUBJECTS}")
    
    base_col = 3 + (day - 1) * len(SUBJECTS)
    return base_col + SUBJECTS.index(subject)

def save_name_mapping(user_id, name):
    """Save user ID and name mapping to file"""
    try:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(NAME_FILE) if os.path.dirname(NAME_FILE) else '.', exist_ok=True)
        
        # Check if mapping already exists
        existing_mappings = load_name_mapping()
        if int(user_id) not in existing_mappings:
            with open(NAME_FILE, "a", encoding='utf-8') as f:
                f.write(f"{user_id},{name}\n")
            print(f"Saved mapping: {user_id} -> {name}")
        else:
            print(f"Mapping for {user_id} already exists")
    except Exception as e:
        print(f"Error saving name mapping: {e}")

def load_name_mapping():
    """Load user ID and name mappings from file"""
    if not os.path.exists(NAME_FILE):
        return {}
    
    try:
        mappings = {}
        with open(NAME_FILE, "r", encoding='utf-8') as f:
            lines = f.readlines()
        
        for line in lines:
            line = line.strip()
            if line and ',' in line:
                parts = line.split(",", 1)  # Split only on first comma
                if len(parts) == 2:
                    mappings[int(parts[0])] = parts[1]
        
        print(f"Loaded {len(mappings)} name mappings")
        return mappings
    except Exception as e:
        print(f"Error loading name mappings: {e}")
        return {}

def collect_faces(face_id, name, save_dir='face_data'):
    """Collect face samples for training"""
    try:
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        face_id = int(face_id)
        save_name_mapping(face_id, name)

        cam = cv2.VideoCapture(0)
        if not cam.isOpened():
            st.error("Cannot access camera")
            return

        detector = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        count = 0
        max_samples = 30

        stframe = st.empty()
        status_text = st.empty()

        while count < max_samples:
            ret, img = cam.read()
            if not ret:
                st.error("Failed to capture image from camera")
                break
                
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, minSize=(50, 50))

            for (x, y, w, h) in faces:
                count += 1
                face_img = gray[y:y+h, x:x+w]
                face_path = f"{save_dir}/User.{face_id}.{count}.jpg"
                cv2.imwrite(face_path, face_img)
                
                # Draw rectangle and text
                cv2.rectangle(img, (x, y), (x+w, y+h), (255, 0, 0), 2)
                cv2.putText(img, f"Samples: {count}/{max_samples}", (x, y-10), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

            # Display image
            stframe.image(cv2.cvtColor(img, cv2.COLOR_BGR2RGB), channels="RGB")
            status_text.text(f"Collecting samples: {count}/{max_samples}")
            
            if count >= max_samples:
                break

        cam.release()
        st.success(f"Face collection complete! Collected {count} samples for {name}")
        
    except Exception as e:
        st.error(f"Error during face collection: {e}")

def train_model(data_path='face_data', model_path='trainer.yml'):
    """Train the face recognition model"""
    try:
        # Check if opencv-contrib-python is available
        try:
            recognizer = cv2.face.LBPHFaceRecognizer_create()
        except AttributeError:
            st.error("cv2.face not available. Please install opencv-contrib-python: pip install opencv-contrib-python")
            return

        def get_images_and_labels(path):
            if not os.path.exists(path):
                return [], []
                
            image_paths = [os.path.join(path, f) for f in os.listdir(path) 
                          if f.startswith("User.") and f.endswith(".jpg")]
            
            face_samples, ids = [], []
            
            for image_path in image_paths:
                try:
                    # Extract ID from filename
                    filename = os.path.split(image_path)[-1]
                    user_id = int(filename.split('.')[1])
                    
                    # Read image
                    gray_img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
                    if gray_img is not None:
                        face_samples.append(gray_img)
                        ids.append(user_id)
                except Exception as e:
                    print(f"Error processing {image_path}: {e}")
                    continue
                    
            return face_samples, ids

        faces, ids = get_images_and_labels(data_path)
        
        if not faces:
            st.error("No face data found to train. Please collect face data first.")
            return

        st.info(f"Training with {len(faces)} face samples...")
        recognizer.train(faces, np.array(ids))
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(model_path) if os.path.dirname(model_path) else '.', exist_ok=True)
        recognizer.write(model_path)
        
        st.success(f"Training complete! Model saved with {len(set(ids))} unique faces.")
        
    except Exception as e:
        st.error(f"Error during training: {e}")

def ensure_excel_structure():
    """Ensure Excel file and directory structure exists"""
    try:
        # Create directory
        excel_dir = os.path.dirname(EXCEL_FILE)
        if excel_dir and not os.path.exists(excel_dir):
            os.makedirs(excel_dir)
            print(f"Created directory: {excel_dir}")

        # Load or create workbook
        if os.path.exists(EXCEL_FILE):
            try:
                wb = load_workbook(EXCEL_FILE)
                sheet = wb.active
                print("Loaded existing Excel file")
            except Exception as e:
                print(f"Error loading existing Excel file: {e}")
                wb = Workbook()
                sheet = wb.active
                print("Created new workbook due to loading error")
        else:
            wb = Workbook()
            sheet = wb.active
            print("Created new Excel file")

        # Check if sheet needs initialization
        if sheet.cell(row=1, column=1).value != "ID":
            print("Initializing sheet structure...")
            if not initialize_sheet(sheet):
                raise Exception("Failed to initialize sheet")

        return wb, sheet
    except Exception as e:
        print(f"Error ensuring Excel structure: {e}")
        raise

def recognize_and_mark(subject):
    """Recognize faces and mark attendance"""
    try:
        # Check if model exists
        model_path = 'trainer.yml'
        if not os.path.exists(model_path):
            st.error("Model not found! Please train the model first.")
            return

        # Check if opencv-contrib-python is available
        try:
            recognizer = cv2.face.LBPHFaceRecognizer_create()
        except AttributeError:
            st.error("cv2.face not available. Please install opencv-contrib-python")
            return

        # Get current day
        today = datetime.now()
        current_day = today.day
        
        if current_day > NUM_DAYS:
            st.error(f"Current day ({current_day}) is beyond the attendance period (1-{NUM_DAYS})")
            return

        # Load model and cascade
        recognizer.read(model_path)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        name_map = load_name_mapping()

        if not name_map:
            st.warning("No name mappings found. Please collect face data first.")

        # Ensure Excel structure
        wb, sheet = ensure_excel_structure()

        # Calculate column positions
        col_idx = get_column_for_subject(current_day, subject)
        subject_total_start = 3 + NUM_DAYS * len(SUBJECTS)
        subject_total_map = {subj: subject_total_start + i for i, subj in enumerate(SUBJECTS)}
        total_attended_col = subject_total_start + len(SUBJECTS)
        total_classes_col = total_attended_col + 1
        percent_col = total_classes_col + 1

        print(f"Marking attendance for {subject} on day {current_day} (column {col_idx})")

        # Start camera
        cam = cv2.VideoCapture(0)
        if not cam.isOpened():
            st.error("Cannot access camera")
            return

        attendance = {}
        stframe = st.empty()
        status_text = st.empty()
        
        st.info(f"Starting attendance for {subject}. Show your face to the camera...")

        # Recognition loop
        frame_count = 0
        while len(attendance) < 10 and frame_count < 300:  # Max 10 seconds at 30fps
            ret, img = cam.read()
            if not ret:
                break
                
            frame_count += 1
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, scaleFactor=1.2, minNeighbors=5, minSize=(50, 50))

            for (x, y, w, h) in faces:
                face_img = gray[y:y+h, x:x+w]
                user_id, confidence = recognizer.predict(face_img)
                
                # Lower confidence = better match
                if confidence < 70:  # Adjusted threshold
                    name = name_map.get(user_id, f"Unknown_{user_id}")
                    attendance[user_id] = name
                    label = f"{name} ✅ ({confidence:.1f})"
                    color = (0, 255, 0)  # Green for recognized
                else:
                    label = f"Unknown ({confidence:.1f})"
                    color = (0, 0, 255)  # Red for unknown
                
                cv2.rectangle(img, (x, y), (x+w, y+h), color, 2)
                cv2.putText(img, label, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

            # Display frame
            stframe.image(cv2.cvtColor(img, cv2.COLOR_BGR2RGB), channels="RGB")
            status_text.text(f"Recognized: {len(attendance)} students")

        cam.release()

        if not attendance:
            st.warning("No faces recognized. Please try again.")
            return

        st.info(f"Recognized {len(attendance)} students. Updating Excel file...")

        # Build ID to row mapping
        id_row_map = {}
        for row_idx in range(3, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value is not None:
                try:
                    student_id = int(cell_value)
                    id_row_map[student_id] = row_idx
                except (ValueError, TypeError):
                    continue

        # Mark attendance for recognized students
        for student_id, name in attendance.items():
            if student_id in id_row_map:
                row = id_row_map[student_id]
            else:
                # Add new student
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=student_id)
                sheet.cell(row=row, column=2, value=name)
                id_row_map[student_id] = row
            
            # Mark present
            sheet.cell(row=row, column=col_idx, value="P")
            print(f"Marked {name} (ID: {student_id}) as Present")

        # Update calculations for all students
        max_row = max(sheet.max_row, max(id_row_map.values()) if id_row_map else 3)
        
        for row_idx in range(3, max_row + 1):
            student_id_cell = sheet.cell(row=row_idx, column=1).value
            if student_id_cell is None:
                continue

            # Mark absents for past days (only mark current day if not already marked)
            for day in range(1, current_day + 1):
                for subj in SUBJECTS:
                    col = get_column_for_subject(day, subj)
                    cell_value = sheet.cell(row=row_idx, column=col).value
                    if cell_value is None or cell_value == "":
                        if day < current_day or (day == current_day and subj != subject):
                            sheet.cell(row=row_idx, column=col, value="A")

            # Calculate subject totals
            for subj in SUBJECTS:
                present_count = 0
                for day in range(1, NUM_DAYS + 1):
                    col = get_column_for_subject(day, subj)
                    if sheet.cell(row=row_idx, column=col).value == "P":
                        present_count += 1
                sheet.cell(row=row_idx, column=subject_total_map[subj], value=present_count)

            # Calculate overall totals
            total_present = 0
            for col in range(3, subject_total_start):
                if sheet.cell(row=row_idx, column=col).value == "P":
                    total_present += 1

            total_classes_so_far = current_day * len(SUBJECTS)
            sheet.cell(row=row_idx, column=total_attended_col, value=total_present)
            sheet.cell(row=row_idx, column=total_classes_col, value=total_classes_so_far)
            
            if total_classes_so_far > 0:
                percentage = round((total_present / total_classes_so_far) * 100, 2)
                sheet.cell(row=row_idx, column=percent_col, value=f"{percentage}%")

        # Save Excel file
        try:
            wb.save(EXCEL_FILE)
            st.success(f"✅ Attendance for {subject} on {current_day} {MONTH_NAME} has been recorded successfully!")
            st.info(f"Excel file saved: {EXCEL_FILE}")
            
            # Show attendance summary
            st.subheader("Attendance Summary:")
            for student_id, name in attendance.items():
                st.write(f"• {name} (ID: {student_id}) - Present")
                
        except PermissionError:
            st.error("❌ Cannot save Excel file. Please close the file if it's open in Excel and try again.")
        except Exception as e:
            st.error(f"❌ Error saving Excel file: {e}")

    except Exception as e:
        st.error(f"Error during attendance marking: {e}")
        print(f"Detailed error: {e}")
        

def get_current_month_year():
    """Get current month and year"""
    now = datetime.now()
    return now.strftime("%B"), now.year, now.month

def get_excel_file_path(month=None, year=None):
    """Get Excel file path for specific month/year"""
    if month is None or year is None:
        month, year, _ = get_current_month_year()
    return f"{EXCEL_BASE_DIR}/{month}_{year}_attendance.xlsx"

def get_days_in_month(month_num, year):
    """Get number of days in a specific month"""
    return calendar.monthrange(year, month_num)[1]

def create_monthly_sheet(month, year, month_num):
    """Create a new monthly attendance sheet"""
    try:
        # Create directory if it doesn't exist
        os.makedirs(EXCEL_BASE_DIR, exist_ok=True)
        
        excel_file = get_excel_file_path(month, year)
        days_in_month = get_days_in_month(month_num, year)
        
        # Create new workbook
        wb = Workbook()
        sheet = wb.active
        sheet.title = f"{month} {year} Attendance"
        
        # Initialize sheet with proper headers
        initialize_monthly_sheet(sheet, month, days_in_month)
        
        # Save the workbook
        wb.save(excel_file)
        st.success(f"✅ Created monthly sheet for {month} {year}")
        return wb, sheet
        
    except Exception as e:
        st.error(f"Error creating monthly sheet: {e}")
        return None, None

def initialize_monthly_sheet(sheet, month, num_days):
    """Initialize the monthly Excel sheet with proper headers"""
    try:
        # Clear existing content
        sheet.delete_rows(1, sheet.max_row)
        
        # Set headers with styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Basic headers
        sheet.cell(row=1, column=1, value="ID")
        sheet.cell(row=1, column=2, value="Name")
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        col = 3
        # Date headers for each day
        for day in range(1, num_days + 1):
            sheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(SUBJECTS) - 1)
            sheet.cell(row=1, column=col, value=f"{day} {month}")
            
            # Subject headers for each day
            for i, subject in enumerate(SUBJECTS):
                sheet.cell(row=2, column=col + i, value=subject)
            col += len(SUBJECTS)

        # Total columns for each subject
        for subject in SUBJECTS:
            sheet.cell(row=1, column=col, value=f"Total {subject}")
            sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            col += 1

        # Summary columns
        summary_headers = ["Total Present", "Total Classes", "Attendance %"]
        for header in summary_headers:
            sheet.cell(row=1, column=col, value=header)
            sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            col += 1

        # Apply styling to headers
        for row in sheet.iter_rows(min_row=1, max_row=2):
            for cell in row:
                if cell.value:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                    
        # Set column widths
        sheet.column_dimensions['A'].width = 8   # ID
        sheet.column_dimensions['B'].width = 20  # Name
        for col_idx in range(3, col):
            col_letter = get_column_letter(col_idx)
            sheet.column_dimensions[col_letter].width = 4  # Subject columns
                    
        print(f"Monthly sheet initialized successfully for {month}")
        return True
    except Exception as e:
        print(f"Error initializing monthly sheet: {e}")
        return False

def ensure_monthly_excel_structure(month=None, year=None):
    """Ensure monthly Excel file exists with proper structure"""
    try:
        if month is None or year is None:
            month, year, month_num = get_current_month_year()
        else:
            month_num = datetime.strptime(month, "%B").month
        
        excel_file = get_excel_file_path(month, year)
        
        # Check if file exists
        if os.path.exists(excel_file):
            try:
                wb = load_workbook(excel_file)
                sheet = wb.active
                print(f"Loaded existing Excel file for {month} {year}")
                
                # Check if sheet needs reinitialization
                if sheet.cell(row=1, column=1).value != "ID":
                    days_in_month = get_days_in_month(month_num, year)
                    if initialize_monthly_sheet(sheet, month, days_in_month):
                        wb.save(excel_file)  # Save after initialization
                        print(f"Reinitialized sheet for {month} {year}")
                
                return wb, sheet
                
            except Exception as e:
                print(f"Error loading existing Excel file: {e}")
                # If loading fails, create new file
                return create_monthly_sheet(month, year, month_num)
        else:
            # File doesn't exist, create new one
            print(f"Creating new Excel file for {month} {year}")
            return create_monthly_sheet(month, year, month_num)

    except Exception as e:
        print(f"Error ensuring monthly Excel structure: {e}")
        st.error(f"Failed to create/load Excel file: {e}")
        return None, None

# ------------------- NEW: Results Display Functions -------------------

def load_monthly_attendance_data(month, year):
    """Load attendance data from monthly Excel file"""
    try:
        excel_file = get_excel_file_path(month, year)
        if not os.path.exists(excel_file):
            return None, "File not found"
        
        # Read Excel file
        df = pd.read_excel(excel_file, header=[0, 1])
        return df, None
    except Exception as e:
        return None, str(e)

def display_attendance_summary(month, year):
    """Display attendance summary for a specific month"""
    try:
        excel_file = get_excel_file_path(month, year)
        if not os.path.exists(excel_file):
            st.warning(f"No attendance data found for {month} {year}")
            return
        
        wb = load_workbook(excel_file)
        sheet = wb.active
        
        # Find summary columns
        summary_data = []
        
        # Read data starting from row 3 (after headers)
        for row_idx in range(3, sheet.max_row + 1):
            student_id = sheet.cell(row=row_idx, column=1).value
            student_name = sheet.cell(row=row_idx, column=2).value
            
            if student_id is None:
                continue
                
            # Find summary columns (last 3 columns typically)
            max_col = sheet.max_column
            total_present = sheet.cell(row=row_idx, column=max_col - 2).value or 0
            total_classes = sheet.cell(row=row_idx, column=max_col - 1).value or 0
            percentage = sheet.cell(row=row_idx, column=max_col).value or "0%"
            
            summary_data.append({
                'ID': student_id,
                'Name': student_name,
                'Present': total_present,
                'Total': total_classes,
                'Percentage': percentage
            })
        
        if summary_data:
            # Display as DataFrame
            df = pd.DataFrame(summary_data)
            st.dataframe(df, use_container_width=True)
            
            # Display statistics
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Students", len(summary_data))
            with col2:
                avg_attendance = sum([float(str(row['Percentage']).replace('%', '')) 
                                    for row in summary_data]) / len(summary_data)
                st.metric("Average Attendance", f"{avg_attendance:.1f}%")
            with col3:
                high_attendance = len([row for row in summary_data 
                                     if float(str(row['Percentage']).replace('%', '')) >= 75])
                st.metric("≥75% Attendance", high_attendance)
            with col4:
                low_attendance = len([row for row in summary_data 
                                    if float(str(row['Percentage']).replace('%', '')) < 75])
                st.metric("<75% Attendance", low_attendance)
        else:
            st.info("No attendance data available")
            
    except Exception as e:
        st.error(f"Error displaying attendance summary: {e}")

def display_subject_wise_attendance(month, year):
    """Display subject-wise attendance statistics"""
    try:
        excel_file = get_excel_file_path(month, year)
        if not os.path.exists(excel_file):
            return
        
        wb = load_workbook(excel_file)
        sheet = wb.active
        
        # Find subject total columns
        month_num = datetime.strptime(month, "%B").month
        days_in_month = get_days_in_month(month_num, year)
        subject_total_start = 3 + days_in_month * len(SUBJECTS)
        
        subject_stats = {subject: [] for subject in SUBJECTS}
        
        # Read subject totals for each student
        for row_idx in range(3, sheet.max_row + 1):
            student_id = sheet.cell(row=row_idx, column=1).value
            if student_id is None:
                continue
                
            for i, subject in enumerate(SUBJECTS):
                col = subject_total_start + i
                present_count = sheet.cell(row=row_idx, column=col).value or 0
                subject_stats[subject].append(present_count)
        
        # Display subject statistics
        if any(subject_stats.values()):
            st.subheader("📊 Subject-wise Statistics")
            
            cols = st.columns(len(SUBJECTS))
            for i, subject in enumerate(SUBJECTS):
                with cols[i]:
                    if subject_stats[subject]:
                        avg_present = sum(subject_stats[subject]) / len(subject_stats[subject])
                        total_students = len(subject_stats[subject])
                        st.metric(
                            f"{subject}",
                            f"{avg_present:.1f}",
                            f"avg classes attended"
                        )
            
            # Create subject comparison chart
            chart_data = pd.DataFrame({
                'Subject': SUBJECTS,
                'Average_Attendance': [sum(subject_stats[subject]) / len(subject_stats[subject]) 
                                     if subject_stats[subject] else 0 for subject in SUBJECTS]
            })
            
            st.bar_chart(chart_data.set_index('Subject'))
            
    except Exception as e:
        st.error(f"Error displaying subject-wise attendance: {e}")

# ------------------- MODIFIED: Update recognition function -------------------

def recognize_and_mark(subject, month=None, year=None):
    """Recognize faces and mark attendance for specific month"""
    try:
        # Check if model exists
        model_path = 'trainer.yml'
        if not os.path.exists(model_path):
            st.error("Model not found! Please train the model first.")
            return

        # Check if opencv-contrib-python is available
        try:
            recognizer = cv2.face.LBPHFaceRecognizer_create()
        except AttributeError:
            st.error("cv2.face not available. Please install opencv-contrib-python")
            return

        # Get current date info
        if month is None or year is None:
            month, year, month_num = get_current_month_year()
        else:
            month_num = datetime.strptime(month, "%B").month
            
        current_day = datetime.now().day
        days_in_month = get_days_in_month(month_num, year)
        
        if current_day > days_in_month:
            st.error(f"Current day ({current_day}) is beyond the month period (1-{days_in_month})")
            return

        # Load model and cascade
        recognizer.read(model_path)
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        name_map = load_name_mapping()

        if not name_map:
            st.warning("No name mappings found. Please collect face data first.")
            return

        # Ensure monthly Excel structure
        wb, sheet = ensure_monthly_excel_structure(month, year)
        if wb is None or sheet is None:
            st.error("Failed to create/load monthly Excel file")
            return

        # Calculate column positions for monthly sheet
        col_idx = get_column_for_subject_monthly(current_day, subject, days_in_month)
        subject_total_start = 3 + days_in_month * len(SUBJECTS)
        subject_total_map = {subj: subject_total_start + i for i, subj in enumerate(SUBJECTS)}
        total_attended_col = subject_total_start + len(SUBJECTS)
        total_classes_col = total_attended_col + 1
        percent_col = total_classes_col + 1

        print(f"Marking attendance for {subject} on day {current_day} {month} (column {col_idx})")

        # Start camera and recognition
        cam = cv2.VideoCapture(0)
        if not cam.isOpened():
            st.error("Cannot access camera")
            return

        attendance = {}
        stframe = st.empty()
        status_text = st.empty()
        
        st.info(f"Starting attendance for {subject}. Show your face to the camera...")

        # Recognition loop
        frame_count = 0
        max_frames = 300  # 10 seconds at 30fps
        
        while len(attendance) < 10 and frame_count < max_frames:
            ret, img = cam.read()
            if not ret:
                break
                
            frame_count += 1
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray, scaleFactor=1.2, minNeighbors=5, minSize=(50, 50))

            for (x, y, w, h) in faces:
                face_img = gray[y:y+h, x:x+w]
                user_id, confidence = recognizer.predict(face_img)
                
                if confidence < 70:
                    name = name_map.get(user_id, f"Unknown_{user_id}")
                    attendance[user_id] = name
                    label = f"{name} ✅ ({confidence:.1f})"
                    color = (0, 255, 0)
                else:
                    label = f"Unknown ({confidence:.1f})"
                    color = (0, 0, 255)
                
                cv2.rectangle(img, (x, y), (x+w, y+h), color, 2)
                cv2.putText(img, label, (x, y-10), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

            stframe.image(cv2.cvtColor(img, cv2.COLOR_BGR2RGB), channels="RGB")
            status_text.text(f"Recognized: {len(attendance)} students | Frame: {frame_count}/{max_frames}")

        cam.release()

        if not attendance:
            st.warning("No faces recognized. Please try again.")
            return

        st.info(f"Recognized {len(attendance)} students. Updating monthly Excel file...")

        # Build ID to row mapping
        id_row_map = {}
        for row_idx in range(3, sheet.max_row + 1):
            cell_value = sheet.cell(row=row_idx, column=1).value
            if cell_value is not None:
                try:
                    student_id = int(cell_value)
                    id_row_map[student_id] = row_idx
                except (ValueError, TypeError):
                    continue

        # Mark attendance for recognized students
        for student_id, name in attendance.items():
            if student_id in id_row_map:
                row = id_row_map[student_id]
                print(f"Found existing student {name} at row {row}")
            else:
                # Add new student
                row = sheet.max_row + 1
                sheet.cell(row=row, column=1, value=student_id)
                sheet.cell(row=row, column=2, value=name)
                id_row_map[student_id] = row
                print(f"Added new student {name} at row {row}")
            
            # Mark present for this subject today
            sheet.cell(row=row, column=col_idx, value="P")
            print(f"Marked {name} (ID: {student_id}) as Present in column {col_idx}")

        # Update calculations for all students in the sheet
        max_row = max(sheet.max_row, max(id_row_map.values()) if id_row_map else 3)
        
        for row_idx in range(3, max_row + 1):
            student_id_cell = sheet.cell(row=row_idx, column=1).value
            if student_id_cell is None:
                continue

            try:
                student_id = int(student_id_cell)
            except (ValueError, TypeError):
                continue

            # Mark absents for past days only (don't mark future days)
            for day in range(1, current_day + 1):
                for subj in SUBJECTS:
                    col = get_column_for_subject_monthly(day, subj, days_in_month)
                    cell_value = sheet.cell(row=row_idx, column=col).value
                    
                    # Only mark absent if cell is empty and it's a past day/subject
                    if cell_value is None or cell_value == "":
                        if day < current_day or (day == current_day and subj != subject):
                            sheet.cell(row=row_idx, column=col, value="A")

            # Calculate subject totals
            for subj in SUBJECTS:
                present_count = 0
                for day in range(1, days_in_month + 1):
                    try:
                        col = get_column_for_subject_monthly(day, subj, days_in_month)
                        if sheet.cell(row=row_idx, column=col).value == "P":
                            present_count += 1
                    except ValueError:
                        continue  # Skip invalid days
                
                # Update subject total
                sheet.cell(row=row_idx, column=subject_total_map[subj], value=present_count)

            # Calculate overall totals
            total_present = 0
            total_classes_so_far = 0
            
            for day in range(1, current_day + 1):
                for subj in SUBJECTS:
                    try:
                        col = get_column_for_subject_monthly(day, subj, days_in_month)
                        cell_value = sheet.cell(row=row_idx, column=col).value
                        if cell_value == "P":
                            total_present += 1
                        if cell_value in ["P", "A"]:  # Count both present and absent as classes held
                            total_classes_so_far += 1
                    except ValueError:
                        continue

            # Update totals
            sheet.cell(row=row_idx, column=total_attended_col, value=total_present)
            sheet.cell(row=row_idx, column=total_classes_col, value=total_classes_so_far)
            
            # Calculate percentage
            if total_classes_so_far > 0:
                percentage = round((total_present / total_classes_so_far) * 100, 2)
                sheet.cell(row=row_idx, column=percent_col, value=f"{percentage}%")
            else:
                sheet.cell(row=row_idx, column=percent_col, value="0%")

        # Save Excel file
        excel_file = get_excel_file_path(month, year)
        try:
            wb.save(excel_file)
            st.success(f"✅ Attendance for {subject} on {current_day} {month} has been recorded successfully!")
            st.info(f"Excel file saved: {excel_file}")
            
            # Show attendance summary
            st.subheader("Attendance Summary:")
            for student_id, name in attendance.items():
                st.write(f"• {name} (ID: {student_id}) - Present ✅")
                
            # Show file info
            st.info(f"📁 Data saved to: {excel_file}")
            st.info(f"📊 Total students processed: {len(attendance)}")
                
        except PermissionError:
            st.error("❌ Cannot save Excel file. Please close the file if it's open in Excel and try again.")
        except Exception as e:
            st.error(f"❌ Error saving Excel file: {e}")
            print(f"Save error details: {e}")

    except Exception as e:
        st.error(f"Error during attendance marking: {e}")
        print(f"Recognition error details: {e}")
        import traceback
        traceback.print_exc()

def get_column_for_subject_monthly(day, subject, days_in_month):
    """Get the column number for a specific day and subject in monthly sheet"""
    if day < 1 or day > days_in_month:
        raise ValueError(f"Day must be between 1 and {days_in_month}")
    if subject not in SUBJECTS:
        raise ValueError(f"Subject must be one of {SUBJECTS}")
    
    base_col = 3 + (day - 1) * len(SUBJECTS)
    return base_col + SUBJECTS.index(subject)
# ------------------- Streamlit UI with Tabs -------------------

def main():
    st.set_page_config(
        page_title="Face Recognition Attendance System",
        page_icon="📸",
        layout="wide"
    )
    
    st.title("📸 Face Recognition Attendance System")
    st.markdown("---")

    # Create tabs
    tab1, tab2, tab3, tab4, tab5 = st.tabs([
        "🏠 Home", 
        "👤 Collect Face Data", 
        "🤖 Train Model", 
        "✅ Mark Attendance", 
        "📊 View Results",
        # "⚙️ System Status"
    ])

    with tab1:
        st.header("Welcome to the Face Recognition Attendance System")
        st.write("""
        This system helps track attendance using face recognition technology. 
        Follow these steps to get started:
        
        1. **Collect Face Data**: Register students by capturing their face samples
        2. **Train Model**: Process the collected data to create a recognition model
        3. **Mark Attendance**: Use face recognition to automatically mark attendance
        4. **View Status**: Monitor system status and view registered students
        """)
        
        # Show system status in columns
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📋 System Status")
            
            # Check if face data exists
            if os.path.exists('face_data') and os.listdir('face_data'):
                face_count = len([f for f in os.listdir('face_data') if f.endswith('.jpg')])
                st.success(f"✅ Face data: {face_count} samples")
            else:
                st.error("❌ No face data found")
            
            # Check if model exists
            if os.path.exists('trainer.yml'):
                st.success("✅ Trained model exists")
            else:
                st.error("❌ No trained model found")
            
            # Check if Excel file exists
            if os.path.exists(EXCEL_FILE):
                st.success("✅ Excel attendance file exists")
            else:
                st.error("❌ No Excel attendance file found")
        
        with col2:
            st.subheader("📅 Current Information")
            current_date = datetime.now()
            st.info(f"**Current Date:** {current_date.day} {MONTH_NAME} {current_date.year}")
            st.info(f"**Tracking Period:** 1-{NUM_DAYS} {MONTH_NAME}")
            st.info(f"**Subjects:** {', '.join(SUBJECTS)}")
            
            # Show registered students count
            name_mappings = load_name_mapping()
            st.info(f"**Registered Students:** {len(name_mappings)}")

    with tab2:
        st.header("👤 Collect Face Data")
        st.write("Register a new student by collecting face samples for training.")
        
        col1, col2 = st.columns([2, 1])
        
        with col1:
            user_id = st.text_input("📝 Enter Register Number (numeric only)", placeholder="e.g., 12345")
            user_name = st.text_input("👤 Enter Full Name", placeholder="e.g., John Doe")
            
            st.info("💡 **Instructions:**")
            st.write("- Make sure you're in good lighting")
            st.write("- Look directly at the camera")
            st.write("- The system will collect 30 face samples")
            st.write("- Move your face slightly during collection")
            
        with col2:
            st.markdown("### 📊 Collection Stats")
            if os.path.exists('face_data'):
                face_files = [f for f in os.listdir('face_data') if f.endswith('.jpg')]
                unique_ids = set()
                for f in face_files:
                    try:
                        uid = f.split('.')[1]
                        unique_ids.add(uid)
                    except:
                        pass
                st.metric("Total Samples", len(face_files))
                st.metric("Registered Students", len(unique_ids))
        
        if st.button("🎯 Start Face Collection", type="primary"):
            if user_id.isdigit() and user_name.strip():
                with st.spinner("📸 Collecting face samples..."):
                    collect_faces(user_id, user_name)
            else:
                st.error("⚠️ Please enter a valid numeric ID and non-empty name.")

    with tab3:
        st.header("🤖 Train Face Recognition Model")
        st.write("Process collected face data to create the recognition model.")
        
        # Check if face data exists
        if os.path.exists('face_data') and os.listdir('face_data'):
            face_files = [f for f in os.listdir('face_data') if f.endswith('.jpg')]
            
            col1, col2 = st.columns([2, 1])
            
            with col1:
                st.success(f"✅ Found {len(face_files)} face samples for training")
                
                # Show breakdown by student
                student_counts = {}
                for f in face_files:
                    try:
                        uid = f.split('.')[1]
                        student_counts[uid] = student_counts.get(uid, 0) + 1
                    except:
                        continue
                
                st.write("**Sample distribution:**")
                name_mappings = load_name_mapping()
                for uid, count in student_counts.items():
                    name = name_mappings.get(int(uid), f"Unknown_{uid}")
                    st.write(f"• {name} (ID: {uid}): {count} samples")
            
            with col2:
                st.markdown("### 📈 Training Info")
                st.metric("Students", len(student_counts))
                st.metric("Total Samples", len(face_files))
                
                if os.path.exists('trainer.yml'):
                    st.success("Model exists")
                else:
                    st.warning("No model yet")
            
            st.markdown("---")
            if st.button("🚀 Start Training", type="primary"):
                with st.spinner("🔄 Training model... This may take a few minutes."):
                    train_model()
        else:
            st.warning("⚠️ No face data found. Please collect face data first.")
            st.info("👆 Go to the 'Collect Face Data' tab to register students.")

    with tab4:
        st.header("✅ Mark Attendance")
        st.write("Use face recognition to automatically mark student attendance.")
        
        # Month/Year selection
        col1, col2 = st.columns(2)
        with col1:
            current_month, current_year, _ = get_current_month_year()
            selected_month = st.selectbox("📅 Select Month", 
                                        [datetime(2024, i, 1).strftime("%B") for i in range(1, 13)],
                                        index=datetime.now().month - 1)
        with col2:
            selected_year = st.selectbox("📅 Select Year", 
                                       [2025, 2026], 
                                       index=0 if current_year == 2024 else 0)
        
        # Subject selection
        subject = st.selectbox("📚 Choose Subject", SUBJECTS, index=0)
        
        # Check prerequisites
        if os.path.exists('trainer.yml'):
            if st.button("📹 Start Attendance", type="primary"):
                with st.spinner(f"🔍 Starting attendance for {subject}..."):
                    recognize_and_mark(subject, selected_month, selected_year)
        else:
            st.error("❌ No trained model found. Please train the model first.")

    with tab5:
        st.header("📊 View Attendance Results")
        
        # Month/Year selection for viewing results
        col1, col2 = st.columns(2)
        with col1:
            view_month = st.selectbox("📅 Select Month to View", 
                                    [datetime(2024, i, 1).strftime("%B") for i in range(1, 13)],
                                    index=datetime.now().month - 1,
                                    key="view_month")
        with col2:
            view_year = st.selectbox("📅 Select Year to View", 
                                   [2025, 2026], 
                                   index=0,
                                   key="view_year")
        
        # Display options
        view_option = st.radio("Select View Type", 
                              ["📈 Summary Report", "📊 Subject-wise Analysis", "📋 Detailed Records"])
        
        if st.button("🔍 Load Results"):
            if view_option == "📈 Summary Report":
                st.subheader(f"Summary Report - {view_month} {view_year}")
                display_attendance_summary(view_month, view_year)
                
            elif view_option == "📊 Subject-wise Analysis":
                st.subheader(f"Subject-wise Analysis - {view_month} {view_year}")
                display_subject_wise_attendance(view_month, view_year)
                
            elif view_option == "📋 Detailed Records":
                st.subheader(f"Detailed Records - {view_month} {view_year}")
                excel_file = get_excel_file_path(view_month, view_year)
                if os.path.exists(excel_file):
                    try:
                        df = pd.read_excel(excel_file)
                        st.dataframe(df, use_container_width=True)
                        
                        # Download button
                        with open(excel_file, "rb") as file:
                            st.download_button(
                                label="📥 Download Excel File",
                                data=file.read(),
                                file_name=f"{view_month}_{view_year}_attendance.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                    except Exception as e:
                        st.error(f"Error loading detailed records: {e}")
                else:
                    st.warning(f"No attendance file found for {view_month} {view_year}")

    # with tab6:
    #     st.header("⚙️ System Status")
        
    #     # List all monthly files
    #     st.subheader("📁 Monthly Attendance Files")
    #     if os.path.exists(EXCEL_BASE_DIR):
    #         files = [f for f in os.listdir(EXCEL_BASE_DIR) if f.endswith('.xlsx')]
    #         if files:
    #             for file in sorted(files):
    #                 file_path = os.path.join(EXCEL_BASE_DIR, file)
    #                 file_size = os.path.getsize(file_path)
    #                 st.write(f"📄 {file} ({file_size} bytes)")
    #         else:
    #             st.info("No monthly attendance files found")
    #     else:
    #         st.info("Attendance directory not created yet")
        
    #     # ... (keep existing system status content)

# ------------------- Original Helper Functions (Updated) -------------------

def initialize_sheet(sheet):
    """Initialize the Excel sheet with proper headers"""
    try:
        # Clear existing content
        sheet.delete_rows(1, sheet.max_row)
        
        # Set headers
        sheet.cell(row=1, column=1, value="ID")
        sheet.cell(row=1, column=2, value="Name")
        sheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        sheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        col = 3
        # Date headers
        for day in range(1, NUM_DAYS + 1):
            if col + len(SUBJECTS) - 1 <= sheet.max_column or sheet.max_column == 1:
                sheet.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + len(SUBJECTS) - 1)
            sheet.cell(row=1, column=col, value=f"{day} {MONTH_NAME}")
            
            # Subject headers for each day
            for i, subject in enumerate(SUBJECTS):
                sheet.cell(row=2, column=col + i, value=subject)
            col += len(SUBJECTS)

        # Total columns for each subject
        for subject in SUBJECTS:
            sheet.cell(row=1, column=col, value=f"Total {subject}")
            sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
            col += 1

        # Summary columns
        sheet.cell(row=1, column=col, value="Total Attended Classes")
        sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

        sheet.cell(row=1, column=col, value="Total Classes")
        sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1

        sheet.cell(row=1, column=col, value="Percentage")
        sheet.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

        # Apply alignment
        for row in sheet.iter_rows(min_row=1, max_row=2):
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    
        print("Sheet initialized successfully")
        return True
    except Exception as e:
        print(f"Error initializing sheet: {e}")
        return False

def get_column_for_subject(day, subject):
    """Get the column number for a specific day and subject"""
    if day < 1 or day > NUM_DAYS:
        raise ValueError(f"Day must be between 1 and {NUM_DAYS}")
    if subject not in SUBJECTS:
        raise ValueError(f"Subject must be one of {SUBJECTS}")
    
    base_col = 3 + (day - 1) * len(SUBJECTS)
    return base_col + SUBJECTS.index(subject)

def save_name_mapping(user_id, name):
    """Save user ID and name mapping to file"""
    try:
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(NAME_FILE) if os.path.dirname(NAME_FILE) else '.', exist_ok=True)
        
        # Check if mapping already exists
        existing_mappings = load_name_mapping()
        if int(user_id) not in existing_mappings:
            with open(NAME_FILE, "a", encoding='utf-8') as f:
                f.write(f"{user_id},{name}\n")
            print(f"Saved mapping: {user_id} -> {name}")
        else:
            print(f"Mapping for {user_id} already exists")
    except Exception as e:
        print(f"Error saving name mapping: {e}")

def load_name_mapping():
    """Load user ID and name mappings from file"""
    if not os.path.exists(NAME_FILE):
        return {}
    
    try:
        mappings = {}
        with open(NAME_FILE, "r", encoding='utf-8') as f:
            lines = f.readlines()
        
        for line in lines:
            line = line.strip()
            if line and ',' in line:
                parts = line.split(",", 1)  # Split only on first comma
                if len(parts) == 2:
                    mappings[int(parts[0])] = parts[1]
        
        print(f"Loaded {len(mappings)} name mappings")
        return mappings
    except Exception as e:
        print(f"Error loading name mappings: {e}")
        return {}

def collect_faces(face_id, name, save_dir='face_data'):
    """Collect face samples for training"""
    try:
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)

        face_id = int(face_id)
        save_name_mapping(face_id, name)

        cam = cv2.VideoCapture(0)
        if not cam.isOpened():
            st.error("Cannot access camera")
            return

        detector = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')
        count = 0
        max_samples = 30

        stframe = st.empty()
        status_text = st.empty()

        while count < max_samples:
            ret, img = cam.read()
            if not ret:
                st.error("Failed to capture image from camera")
                break
                
            gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
            faces = detector.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, minSize=(50, 50))

            for (x, y, w, h) in faces:
                count += 1
                face_img = gray[y:y+h, x:x+w]
                face_path = f"{save_dir}/User.{face_id}.{count}.jpg"
                cv2.imwrite(face_path, face_img)
                
                # Draw rectangle and text
                cv2.rectangle(img, (x, y), (x+w, y+h), (255, 0, 0), 2)
                cv2.putText(img, f"Samples: {count}/{max_samples}", (x, y-10), 
                           cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

            # Display image
            stframe.image(cv2.cvtColor(img, cv2.COLOR_BGR2RGB), channels="RGB")
            status_text.text(f"Collecting samples: {count}/{max_samples}")
            
            if count >= max_samples:
                break

        cam.release()
        st.success(f"Face collection complete! Collected {count} samples for {name}")
        
    except Exception as e:
        st.error(f"Error during face collection: {e}")

def train_model(data_path='face_data', model_path='trainer.yml'):
    """Train the face recognition model"""
    try:
        # Check if opencv-contrib-python is available
        try:
            recognizer = cv2.face.LBPHFaceRecognizer_create()
        except AttributeError:
            st.error("cv2.face not available. Please install opencv-contrib-python: pip install opencv-contrib-python")
            return

        def get_images_and_labels(path):
            if not os.path.exists(path):
                return [], []
                
            image_paths = [os.path.join(path, f) for f in os.listdir(path) 
                          if f.startswith("User.") and f.endswith(".jpg")]
            
            face_samples, ids = [], []
            
            for image_path in image_paths:
                try:
                    # Extract ID from filename
                    filename = os.path.split(image_path)[-1]
                    user_id = int(filename.split('.')[1])
                    
                    # Read image
                    gray_img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
                    if gray_img is not None:
                        face_samples.append(gray_img)
                        ids.append(user_id)
                except Exception as e:
                    print(f"Error processing {image_path}: {e}")
                    continue
                    
            return face_samples, ids

        faces, ids = get_images_and_labels(data_path)
        
        if not faces:
            st.error("No face data found to train. Please collect face data first.")
            return

        st.info(f"Training with {len(faces)} face samples...")
        recognizer.train(faces, np.array(ids))
        
        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(model_path) if os.path.dirname(model_path) else '.', exist_ok=True)
        recognizer.write(model_path)
        
        st.success(f"Training complete! Model saved with {len(set(ids))} unique faces.")
        
    except Exception as e:
        st.error(f"Error during training: {e}")

def ensure_excel_structure():
    """Ensure Excel file and directory structure exists"""
    try:
        # Create directory
        excel_dir = os.path.dirname(EXCEL_FILE)
        if excel_dir and not os.path.exists(excel_dir):
            os.makedirs(excel_dir)
            print(f"Created directory: {excel_dir}")

        # Load or create workbook
        if os.path.exists(EXCEL_FILE):
            try:
                wb = load_workbook(EXCEL_FILE)
                sheet = wb.active
                print("Loaded existing Excel file")
            except Exception as e:
                print(f"Error loading existing Excel file: {e}")
                wb = Workbook()
                sheet = wb.active
                print("Created new workbook due to loading error")
        else:
            wb = Workbook()
            sheet = wb.active
            print("Created new Excel file")

        # Check if sheet needs initialization
        if sheet.cell(row=1, column=1).value != "ID":
            print("Initializing sheet structure...")
            if not initialize_sheet(sheet):
                raise Exception("Failed to initialize sheet")

        return wb, sheet
    except Exception as e:
        print(f"Error ensuring Excel structure: {e}")
        raise

if __name__ == "__main__":
    main()